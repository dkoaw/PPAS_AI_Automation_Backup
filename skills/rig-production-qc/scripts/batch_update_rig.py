# -*- coding: utf-8 -*-
import os, sys, json, codecs

try:
    import openpyxl
except ImportError:
    pass

def batch_update_cells(project_name, updates_payload):
    """
    updates_payload format: asset_name|column_name|new_value;asset_name|column_name|new_value
    """
    spreadsheet_path = os.path.join("X:/AI_Automation/Project", project_name, "work", "spreadsheet", u"{}_Rig制作管理表.xlsx".format(project_name))
    if not os.path.exists(spreadsheet_path):
        print("Error: Spreadsheet not found at {}".format(spreadsheet_path))
        sys.exit(1)

    updates = []
    for item in updates_payload.split(';'):
        if not item.strip(): continue
        parts = item.split('|')
        if len(parts) == 3:
            updates.append((parts[0].strip(), parts[1].strip(), parts[2].strip()))

    if not updates: return

    try:
        import subprocess
        subprocess.call(['attrib', '-r', spreadsheet_path.encode('gbk')])
        wb = openpyxl.load_workbook(spreadsheet_path)
        ws = wb.active

        headers = {cell.value: i for i, cell in enumerate(ws[1]) if cell.value}
        
        updated_count = 0
        for row in ws.iter_rows(min_row=2):
            asset_name_cell = row[headers.get(u"资产名称")]
            if not asset_name_cell or not asset_name_cell.value: continue
            asset_name = unicode(asset_name_cell.value)
            
            for up_asset, up_col, up_val in updates:
                if asset_name == unicode(up_asset):
                    col_idx = headers.get(unicode(up_col))
                    if col_idx is not None:
                        row[col_idx].value = up_val
                        updated_count += 1

        wb.save(spreadsheet_path)
        subprocess.call(['attrib', '+r', spreadsheet_path.encode('gbk')])
        print("Batch updated {} cells in Rig QC Table".format(updated_count))
    except Exception as e:
        print("Batch Update Error: {}".format(e))
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3: sys.exit(1)
    batch_update_cells(sys.argv[1], sys.argv[2].decode('utf-8'))