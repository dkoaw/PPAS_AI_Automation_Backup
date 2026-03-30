# -*- coding: utf-8 -*-
import os
import json
import io
import subprocess
import sys
import codecs

try:
    import xlsxwriter
    import openpyxl
except ImportError:
    pass

def run_sg_query(project_name):
    """
    Dedicated Query for Rig QC. Cares about rigMaster and upstream tasks.
    """
    query_in = {
        "action": "raw_query",
        "entity_type": "Task",
        "filters": [
            ["project.Project.name", "is", project_name],
            ["entity.Asset.sg_asset_type", "in", ["chr", "prp", "env"]],
            ["content", "in", ["modMaster", "uvMaster", "texMaster", "rigMaster"]]
        ],
        "fields": ["entity", "content", "sg_status_list", "entity.Asset.sg_asset_type"]
    }
    
    ts = str(int(time.time() * 1000))
    tmp_in = "X:/AI_Automation/.gemini/tmp/rig_sg_query_in_{}.json".format(ts)
    tmp_out = "X:/AI_Automation/.gemini/tmp/rig_sg_query_out_{}.json".format(ts)
    
    with io.open(tmp_in, 'w', encoding='utf-8') as f:
        f.write(unicode(json.dumps(query_in, ensure_ascii=False)))
        
    sg_script = "X:/AI_Automation/.gemini/skills/sg-data-reader/scripts/sg_query.py"
    blender_path = "C:/Program Files/Blender Foundation/Blender 5.0/blender.exe"
    
    env = {str(k): str(v) for k, v in os.environ.items()}
    env["SG_QUERY_IN"] = str(tmp_in)
    env["SG_QUERY_OUT"] = str(tmp_out)
    
    subprocess.call([blender_path, "-b", "-P", sg_script], env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    
    if os.path.exists(tmp_out):
        with io.open(tmp_out, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('data', [])
    return []

def manage_rig_table(project_name):
    spreadsheet_dir = os.path.join("X:/AI_Automation/Project", project_name, "work", "spreadsheet")
    if not os.path.exists(spreadsheet_dir):
        os.makedirs(spreadsheet_dir)
        
    file_path_xlsx = os.path.join(spreadsheet_dir, u"{}_Rig制作管理表.xlsx".format(project_name))
    
    if os.path.exists(file_path_xlsx):
        subprocess.call(['attrib', '-r', file_path_xlsx.encode('gbk')])

    print(u"Updating Rig QC Spreadsheet: {}...".format(file_path_xlsx))
    sg_tasks = run_sg_query(project_name)
    
    sg_map = {}
    for task in sg_tasks:
        asset_name = task['entity']['name']
        task_name = task['content']
        status = task['sg_status_list']
        asset_type = task.get('entity.Asset.sg_asset_type', 'unknown')
        
        if asset_name not in sg_map:
            sg_map[asset_name] = {"type": asset_type}
        sg_map[asset_name][task_name] = status

    headers = [u"资产类型", u"资产名称", u"modMaster", u"uvMaster", u"texMaster", u"rigMaster", u"rigQC"]
    status_list = [u"wtg", u"rdy", u"ip", u"fin", u"apr", u"pub", u"tmpub", u"rep", u"rtk", u"hld", u"omt"]
    
    existing_data = []
    if os.path.exists(file_path_xlsx):
        try:
            wb = openpyxl.load_workbook(file_path_xlsx)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                file_headers = [unicode(h) if h else u"" for h in rows[0]]
                for r in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(r):
                        if i < len(file_headers) and file_headers[i]:
                            row_dict[file_headers[i]] = val
                    existing_data.append(row_dict)
        except Exception as e:
            print("Warning: Failed to read existing XLSX: {}".format(e))

    final_rows = []
    processed_assets = set()
    
    for row in existing_data:
        asset_name = row.get(u"资产名称")
        if asset_name in sg_map:
            row[u"modMaster"] = sg_map[asset_name].get("modMaster", row.get(u"modMaster", u""))
            row[u"uvMaster"] = sg_map[asset_name].get("uvMaster", row.get(u"uvMaster", u""))
            row[u"texMaster"] = sg_map[asset_name].get("texMaster", row.get(u"texMaster", u""))
            row[u"rigMaster"] = sg_map[asset_name].get("rigMaster", row.get(u"rigMaster", u""))
            processed_assets.add(asset_name)
        final_rows.append(row)
        
    for asset_name in sorted(sg_map.keys()):
        if asset_name not in processed_assets:
            info = sg_map[asset_name]
            new_row = {h: u"" for h in headers}
            new_row[u"资产类型"] = info["type"]
            new_row[u"资产名称"] = asset_name
            new_row[u"modMaster"] = info.get("modMaster", u"")
            new_row[u"uvMaster"] = info.get("uvMaster", u"")
            new_row[u"texMaster"] = info.get("texMaster", u"")
            new_row[u"rigMaster"] = info.get("rigMaster", u"")
            new_row[u"rigQC"] = u"wtg"
            final_rows.append(new_row)

    try:
        workbook = xlsxwriter.Workbook(file_path_xlsx)
        worksheet = workbook.add_worksheet("Rig_QC")
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#E2EFDA', 'border': 1, 'align': 'center'})
        data_format = workbook.add_format({'border': 1})
        
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
            
        for row_num, row_data in enumerate(final_rows):
            excel_row = row_num + 1
            for col_num, header in enumerate(headers):
                val = row_data.get(header, u"")
                worksheet.write(excel_row, col_num, val, data_format)
                
                if header in [u"modMaster", u"uvMaster", u"texMaster", u"rigMaster", u"rigQC"]:
                    worksheet.data_validation(excel_row, col_num, excel_row, col_num, {
                        'validate': 'list', 'source': status_list
                    })
        
        worksheet.set_column(0, 0, 10)
        worksheet.set_column(1, 1, 30)
        worksheet.set_column(2, 6, 12)
        workbook.close()
        
        subprocess.call(['attrib', '+r', file_path_xlsx.encode('gbk')])
        print("Rig QC Table updated and Locked: {}".format(file_path_xlsx.encode('gbk')))
    except Exception as e:
        print("Error writing XLSX: {}".format(e))
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2: sys.exit(1)
    manage_rig_table(sys.argv[1])