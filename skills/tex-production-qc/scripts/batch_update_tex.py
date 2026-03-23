# -*- coding: utf-8 -*-
import os, sys, codecs
import openpyxl

def batch_update_tex(spreadsheet_path, payload):
    """
    Isolated batch updater for Tex Management Spreadsheet.
    payload: "AssetName|ColumnName|Value;..."
    """
    if isinstance(spreadsheet_path, str):
        spreadsheet_path = spreadsheet_path.decode(sys.getfilesystemencoding())
    if isinstance(payload, str):
        payload = payload.decode(sys.getfilesystemencoding())

    if not os.path.exists(spreadsheet_path):
        print("Error: Spreadsheet not found: " + spreadsheet_path)
        return

    # Software Lock: Unlock before writing
    import subprocess
    subprocess.call(['attrib', '-r', spreadsheet_path.encode('gbk')])

    try:
        wb = openpyxl.load_workbook(spreadsheet_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [unicode(h) for h in rows[0]]
        
        # Build mapping
        name_col = headers.index(u"资产名称")
        updates = payload.split(';')
        
        for up in updates:
            parts = up.split('|')
            if len(parts) < 3: continue
            asset_name, col_name, val = parts[0], parts[1], parts[2]
            
            if col_name not in headers:
                print("Warning: Column {} not in table.".format(col_name))
                continue
                
            target_col = headers.index(col_name)
            
            # Find asset row
            found = False
            for row_idx, row_data in enumerate(rows[1:], start=2):
                if unicode(row_data[name_col]) == asset_name:
                    ws.cell(row=row_idx, column=target_col+1).value = val
                    found = True
                    break
            if not found:
                print("Warning: Asset {} not found in table.".format(asset_name))
        
        wb.save(spreadsheet_path)
        # Software Lock: Re-lock after writing
        subprocess.call(['attrib', '+r', spreadsheet_path.encode('gbk')])
        print("Tex Spreadsheet updated and Locked successfully.")
        
    except Exception as e:
        print("Error updating spreadsheet: " + str(e))
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3: sys.exit(1)
    batch_update_tex(sys.argv[1], sys.argv[2])
