# -*- coding: utf-8 -*-
import os
import json
import io
import subprocess
import sys
import codecs

# Import Excel libraries
try:
    import xlsxwriter
    import openpyxl
except ImportError:
    pass

def run_sg_query(project_name):
    """
    Dedicated Query for Tex Production
    Only cares about texMaster related tasks.
    """
    query_in = {
        "action": "raw_query",
        "entity_type": "Task",
        "filters": [
            ["project.Project.name", "is", project_name],
            ["entity.Asset.sg_asset_type", "in", ["chr", "prp"]],
            ["content", "is", "texMaster"]
        ],
        "fields": ["entity", "content", "sg_status_list", "entity.Asset.sg_asset_type"]
    }
    
    tmp_in = "X:/AI_Automation/.gemini/tmp/tex_sg_query_in.json"
    tmp_out = "X:/AI_Automation/.gemini/tmp/tex_sg_query_out.json"
    
    with io.open(tmp_in, 'w', encoding='utf-8') as f:
        f.write(unicode(json.dumps(query_in, ensure_ascii=False)))
        
    sg_script = "X:/AI_Automation/.gemini/skills/sg-data-reader/scripts/sg_query.py"
    blender_path = "C:/Program Files/Blender Foundation/Blender 5.0/blender.exe"
    
    env = {str(k): str(v) for k, v in os.environ.items()}
    env["SG_QUERY_IN"] = str(tmp_in)
    env["SG_QUERY_OUT"] = str(tmp_out)
    
    subprocess.call([blender_path, "-b", "-P", sg_script], env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    
    if os.path.exists(tmp_out):
        with io.open(tmp_out, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('data', [])
    return []

def manage_tex_table(project_name):
    spreadsheet_dir = os.path.join("X:/AI_Automation/Project", project_name, "work", "spreadsheet")
    if not os.path.exists(spreadsheet_dir):
        os.makedirs(spreadsheet_dir)
        
    file_path_xlsx = os.path.join(spreadsheet_dir, u"{}_Tex制作管理表.xlsx".format(project_name))
    
    # Software Lock: Unlock before writing
    if os.path.exists(file_path_xlsx):
        subprocess.call(['attrib', '-r', file_path_xlsx.encode('gbk')])

    print(u"Updating Tex Production Spreadsheet: {}...".format(file_path_xlsx))
    sg_tasks = run_sg_query(project_name)
    
    sg_map = {}
    for task in sg_tasks:
        asset_name = task['entity']['name']
        status = task['sg_status_list']
        asset_type = task.get('entity.Asset.sg_asset_type', 'unknown')
        sg_map[asset_name] = {"type": asset_type, "texMaster": status}

    # Isolated Headers
    headers = [u"资产类型", u"资产名称", u"texMaster", u"texQC"]
    status_list = [u"wtg", u"rdy", u"ip", u"fin", u"apr", u"pub", u"tmpub", u"rtk", u"hld", u"omt"]
    
    existing_data = []
    if os.path.exists(file_path_xlsx):
        try:
            wb = openpyxl.load_workbook(file_path_xlsx)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                file_headers = [unicode(h) if h else u"" for h in rows[0]]
                for r in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(r):
                        if i < len(file_headers) and file_headers[i]:
                            row_dict[file_headers[i]] = val
                    existing_data.append(row_dict)
        except Exception as e:
            print("Warning: Failed to read existing XLSX: {}".format(e))

    final_rows = []
    processed_assets = set()
    
    for row in existing_data:
        asset_name = row.get(u"资产名称")
        if asset_name in sg_map:
            row[u"texMaster"] = sg_map[asset_name].get("texMaster", row.get(u"texMaster", u""))
            processed_assets.add(asset_name)
        final_rows.append(row)
        
    for asset_name in sorted(sg_map.keys()):
        if asset_name not in processed_assets:
            info = sg_map[asset_name]
            new_row = {h: u"" for h in headers}
            new_row[u"资产类型"] = info["type"]
            new_row[u"资产名称"] = asset_name
            new_row[u"texMaster"] = info.get("texMaster", u"")
            new_row[u"texQC"] = u"wtg"
            final_rows.append(new_row)

    try:
        workbook = xlsxwriter.Workbook(file_path_xlsx)
        worksheet = workbook.add_worksheet("Production_QC")
        
        header_format = workbook.add_format({'bold': True, 'bg_color': '#DCE6F1', 'border': 1, 'align': 'center'})
        data_format = workbook.add_format({'border': 1})
        
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
            
        for row_num, row_data in enumerate(final_rows):
            excel_row = row_num + 1
            for col_num, header in enumerate(headers):
                val = row_data.get(header, u"")
                worksheet.write(excel_row, col_num, val, data_format)
                
                if header in [u"texMaster", u"texQC"]:
                    worksheet.data_validation(excel_row, col_num, excel_row, col_num, {
                        'validate': 'list', 'source': status_list
                    })
        
        worksheet.set_column(0, 0, 10)
        worksheet.set_column(1, 1, 25)
        worksheet.set_column(2, 3, 15)
        workbook.close()
        # Software Lock: Re-lock after writing
        subprocess.call(['attrib', '+r', file_path_xlsx.encode('gbk')])
        print("Tex Management Table updated and Locked: {}".format(file_path_xlsx.encode('gbk')))
    except Exception as e:
        print("Error writing XLSX: {}".format(e))
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2: sys.exit(1)
    manage_tex_table(sys.argv[1])
