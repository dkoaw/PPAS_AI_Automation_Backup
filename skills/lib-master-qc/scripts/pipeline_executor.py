# -*- coding: utf-8 -*-
import os
import json
import io
import sys
import codecs
import time
import shutil
import subprocess

# ----------------- 核心组件配置 -----------------
SKILLS_DIR = r"X:\AI_Automation\.gemini\skills"
TMP_DIR = r"X:\AI_Automation\.gemini\tmp"

BLENDER_PATH = r"C:\Program Files\Blender Foundation\Blender 5.0\blender.exe"
FIXER_SCRIPT = os.path.join(SKILLS_DIR, "character-asset-name-fixer", "scripts", "rename_fixer.py")
QC_SCRIPT = os.path.join(SKILLS_DIR, "blender-character-qc", "scripts", "qc_executor.py")
INFO_EXPORTER = os.path.join(SKILLS_DIR, "blender-asset-info-exporter", "scripts", "export_info.py")
SPREADSHEET_MGR = os.path.join(SKILLS_DIR, "asset-spreadsheet-manager", "scripts", "manage_spreadsheet.py")
SCREENSHOT_SCRIPT = os.path.join(SKILLS_DIR, "blender-screenshot", "scripts", "capture_shot.py")
COMPARATOR = os.path.join(SKILLS_DIR, "blender-asset-info-comparator", "scripts", "compare_asset_data.py")

class AssetResult:
    def __init__(self, asset_data):
        self.name = asset_data[u'资产名称']
        self.type = asset_data.get(u'资产类型', 'chr')
        self.tex_status = asset_data.get(u'texMaster', '')
        self.rig_status = asset_data.get(u'rigMaster', '')
        self.lgt_status = asset_data.get(u'灯光文件制作', '')
        self.lib_master_status = asset_data.get(u'libMaster', '')
        self.qc1_existing = asset_data.get(u'QC_step_1', '')
        self.librig_existing = asset_data.get(u'libRig', '')
        
        self.step1_res = "SKIP"
        self.step1_msg = ""
        self.qc1_fail_summary = []
        self.rig_res = "SKIP"
        self.rig_msg = ""
        self.step2_res = "SKIP"
        self.step2_msg = ""
        self.errors = []
        self.screenshot_path = "" 

def log(msg):
    if isinstance(msg, str): msg = msg.decode('utf-8')
    print(u"[lib-master-qc] " + msg)

def get_spreadsheet_data(project_name):
    subprocess.call(["python", SPREADSHEET_MGR, project_name])
    subprocess.call(["python", os.path.join(TMP_DIR, "read_sheet.py"), project_name])
    data_path = os.path.join(TMP_DIR, "spreadsheet_data.json")
    if os.path.exists(data_path):
        with codecs.open(data_path, 'r', 'utf-8') as f: return json.load(f)
    return []

def get_latest_file(directory, pattern):
    import glob
    search_path = os.path.join(directory, pattern)
    files = glob.glob(search_path)
    if not files: return None
    return max(files, key=os.path.getmtime)

def execute_pipeline(project_name, target_asset=None):
    raw_assets = get_spreadsheet_data(project_name)
    if not raw_assets: return
    results = []
    for asset_data in raw_assets:
        if asset_data.get(u'资产类型') != u'chr': continue
        name = asset_data.get(u'资产名称')
        if target_asset and name != target_asset: continue
        
        res = AssetResult(asset_data)
        results.append(res)
        try:
            # --- 严格执行 Excel 规则判定 ---
            # 1. 检查 QC_step_1 条件
            if res.tex_status in ['pub', 'tmpub'] and res.qc1_existing != res.tex_status:
                run_step_1_logic(res, project_name, asset_data)
            else:
                res.step1_msg = u"不符合QC1执行条件 (状态未变或未发布)"
            
            # 2. 检查 libRig 条件
            if ((res.step1_res == "PASS") or (res.qc1_existing == res.tex_status)) and \
               (res.rig_status in ['pub', 'tmpub']) and (res.librig_existing != res.rig_status):
                run_lib_rig_logic(res, project_name, asset_data)
                
        except Exception as e:
            res.errors.append(unicode(str(e)))
    finalize_all(project_name, results)

def run_step_1_logic(res, project_name, asset_data):
    log(u"[{}] Starting QC_step_1...".format(res.name))
    src_dir = os.path.join(r"X:\Project", project_name, r"pub\assets", res.type, res.name, r"tex\texMaster")
    latest = get_latest_file(src_dir, "ysj_{}_{}_tex_texMaster_v*.blend".format(res.type, res.name))
    if not latest: res.step1_res = "FAIL"; res.step1_msg = u"找不到源文件"; return
    
    s1_dir = os.path.join(r"X:\AI_Automation\Project", project_name, r"work\assets_lib", res.type, res.name, "QC_step_1")
    if not os.path.exists(s1_dir): os.makedirs(s1_dir)
    dest_path = os.path.join(s1_dir, u"{}_{}_{}_lib_libMaster.blend".format(project_name, res.type, res.name))
    shutil.copy2(latest, dest_path)
    
    # 执行洗稿 (Fixer V11)
    env = {str(k): str(v) for k, v in os.environ.items()}
    subprocess.call([BLENDER_PATH, "-b", dest_path, "-P", FIXER_SCRIPT], env=env)
    fixed_path = dest_path.replace(".blend", "_fixed.blend")
    
    # 截图
    log(u"[{}] Capturing Screenshot...".format(res.name))
    env["BLENDER_SHOT_OUT"] = str(s1_dir); env["BLENDER_ASSET_NAME"] = str(res.name)
    subprocess.call([BLENDER_PATH, fixed_path, "--python", SCREENSHOT_SCRIPT], env=env)
    
    # 质检
    qc_out = os.path.join(s1_dir, "qc_out.json")
    env["QC_STEP_NAME"] = "tex"; env["QC_OUT_PATH"] = str(qc_out)
    subprocess.call([BLENDER_PATH, "-b", fixed_path, "-P", QC_SCRIPT], env=env)
    
    try:
        with io.open(qc_out, 'r', encoding='utf-8') as f:
            data = json.load(f)
            failed_items = [i for i in data if i.get("status") == "FAIL"]
            if not failed_items:
                res.step1_res = "PASS"
                env["EXTRACT_INFO_OUT"] = str(fixed_path.replace(".blend", ".json"))
                subprocess.call([BLENDER_PATH, "-b", fixed_path, "-P", INFO_EXPORTER], env=env)
            else: 
                res.step1_res = "FAIL"
                res.qc1_fail_summary = [u"{} ({})".format(i['check'], u", ".join(i['issues'])) for i in failed_items]
    except: res.step1_res = "FAIL"

def run_lib_rig_logic(res, project_name, asset_data):
    log(u"[{}] Starting lib-rig...".format(res.name))
    rig_inf = get_latest_file(os.path.join(r"X:\Project", project_name, r"pub\assets", res.type, res.name, r"rig\rigMaster\.info"), "ysj_*.json")
    lib_inf = os.path.join(r"X:\AI_Automation\Project", project_name, r"work\assets_lib", res.type, res.name, "QC_step_1", u"{}_{}_{}_lib_libMaster_fixed.json".format(project_name, res.type, res.name))
    if not (rig_inf and os.path.exists(lib_inf)): res.rig_res = "FAIL"; return
    out_md = os.path.join(os.path.dirname(lib_inf), "lib_rig_diff.md")
    subprocess.call(["python", COMPARATOR, rig_inf, lib_inf, out_md])
    with codecs.open(out_md, 'r', 'utf-8') as f:
        if u"完！全！一！致！" in f.read(): res.rig_res = "PASS"
        else:
            res.rig_res = "FAIL"
            subprocess.call([BLENDER_PATH, "-b", lib_inf.replace(".json", ".blend"), "-P", os.path.join(SKILLS_DIR, "blender-export-abc", "scripts", "export_abc.py")])

def run_step_2_logic(res, project_name, asset_data):
    pass

def finalize_all(project_name, results):
    log(u"Finalizing Report...")
    ups = []; rpt = [u"# 🚀 lib 自动化入库总表", u"时间: {}\n".format(time.strftime("%Y-%m-%d %H:%M:%S"))]
    active_results = [r for r in results if r.step1_res != "SKIP" or r.rig_res != "SKIP"]
    
    if not active_results:
        rpt.append(u"### ℹ️ 本次无资产符合入库流转条件 (全部跳过)。")
    else:
        for res in active_results:
            b = [u"## 📦 资产: `{}`".format(res.name)]
            if res.step1_res == "PASS": 
                ups.append((res.name, u"QC_step_1", res.tex_status))
                b.append(u"### 📝 QC1: ✅ PASS")
            elif res.step1_res == "FAIL": 
                ups.append((res.name, u"QC_step_1", u"rtk"))
                b.append(u"### 📝 QC1: ❌ FAIL (打回)")
                for msg in res.qc1_fail_summary: b.append(u"  * {}".format(msg))
            
            if res.rig_res == "PASS": 
                ups.append((res.name, u"libRig", res.rig_status))
                b.append(u"### 💍 Rig 比对: ✅ PASS")
            elif res.rig_res == "FAIL": 
                ups.append((res.name, u"libRig", u"rtk"))
                b.append(u"### 💍 Rig 比对: ⚠️ 不一致 (已导出 ABC)")
            
            rpt.append(u"\n".join(b) + u"\n---")
        
    apply_batch(project_name, ups)
    p = os.path.join(r"X:\AI_Automation\Project", project_name, "report", "lib_qc_final_{}.md".format(time.strftime("%Y_%m_%d_%H%M")))
    with codecs.open(p, 'w', 'utf-8') as f: f.write(u"\n".join(rpt))
    subprocess.call('start notepad++ "{}"'.format(p.encode('gbk')), shell=True)

def apply_batch(project_name, ups):
    if not ups: return
    bs = os.path.join(TMP_DIR, "batch_update_cells.py")
    with codecs.open(bs, 'w', 'utf-8') as f:
        f.write(u'''# -*- coding: utf-8 -*-
import openpyxl, os, subprocess, sys
pn = sys.argv[1].decode('utf-8')
fp = u"X:/AI_Automation/Project/{0}/work/spreadsheet/{0}_资产入库管理表.xlsx".format(pn)
subprocess.call('attrib -r "{}"'.format(fp.encode('gbk')), shell=True)
wb = openpyxl.load_workbook(fp); ws = wb.active; hs = {c.value: i+1 for i,c in enumerate(ws[1]) if c.value}
for e in sys.argv[2].decode('utf-8').split(';'):
    if not e: continue
    a, c, v = e.split('|')
    for r in range(2, ws.max_row+1):
        ca = ws.cell(row=r, column=2).value
        if (ca if isinstance(ca, unicode) else unicode(ca) if ca else u"") == a:
            if c in hs: ws.cell(row=r, column=hs[c]).value = v
            break
wb.save(fp); subprocess.call('attrib +r "{}"'.format(fp.encode('gbk')), shell=True)
''')
    pl = u";".join([u"{}|{}|{}".format(a, c, v) for a, c, v in ups])
    subprocess.call(["python", bs, project_name.encode('utf-8'), pl.encode('utf-8')])

if __name__ == "__main__":
    proj = sys.argv[1] if len(sys.argv) > 1 else "ysj"
    tgt = sys.argv[2] if len(sys.argv) > 2 else None
    execute_pipeline(proj, tgt)
