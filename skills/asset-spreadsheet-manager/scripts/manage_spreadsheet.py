# -*- coding: utf-8 -*-
import os, json, csv, sys, io, subprocess, codecs

# Import Excel libraries
try:
    import xlsxwriter
    import openpyxl
except ImportError: pass

SKILLS_DIR = r"X:\AI_Automation\Project\..\.gemini\skills"
SG_UPDATE_SCRIPT = os.path.join(SKILLS_DIR, "sg-data-reader", "scripts", "sg_batch_update.py")
BLENDER_PATH = r"C:\Program Files\Blender Foundation\Blender 5.0\blender.exe"

def run_sg_query(project_name):
    query_in = {
        "action": "raw_query", "entity_type": "Task",
        "filters": [
            ["project.Project.name", "is", project_name],
            ["entity.Asset.sg_asset_type", "in", ["chr", "prp", "env"]],
            ["content", "in", ["texMaster", "rigMaster", "facialTex", "lightMap", "libRig", "libMaster"]]
        ],
        "fields": ["entity", "content", "sg_status_list", "entity.Asset.sg_asset_type"]
    }
    tmp_in = "X:/AI_Automation/.gemini/tmp/sg_batch_query_in.json"
    tmp_out = "X:/AI_Automation/.gemini/tmp/sg_batch_query_out.json"
    with io.open(tmp_in, 'w', encoding='utf-8') as f: f.write(unicode(json.dumps(query_in, ensure_ascii=False)))
    sg_script = "X:/AI_Automation/.gemini/skills/sg-data-reader/scripts/sg_query.py"
    env = {str(k): str(v) for k, v in os.environ.items()}
    env["SG_QUERY_IN"] = str(tmp_in); env["SG_QUERY_OUT"] = str(tmp_out)
    subprocess.call([BLENDER_PATH, "-b", "-P", sg_script], env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if os.path.exists(tmp_out):
        with io.open(tmp_out, 'r', encoding='utf-8') as f: data = json.load(f)
        return data.get('data', [])
    return []

def manage_table(project_name):
    spreadsheet_dir = os.path.join("X:/AI_Automation/Project", project_name, "work", "spreadsheet")
    file_path_xlsx = os.path.join(spreadsheet_dir, u"{}_资产入库管理表.xlsx".format(project_name))
    sg_tasks = run_sg_query(project_name)
    sg_map = {}
    for task in sg_tasks:
        asset_name = task['entity']['name']; task_name = task['content']; status = task['sg_status_list']
        asset_type = task.get('entity.Asset.sg_asset_type', 'unknown')
        if asset_name not in sg_map: sg_map[asset_name] = {"type": asset_type}
        sg_map[asset_name][task_name] = status

    headers = [u"资产类型", u"资产名称", u"texMaster", u"rigMaster", u"facialTex", u"lightMap", u"QC_step_1", u"灯光文件制作", u"制作人员", u"libRig", u"QC_step_2", u"libMaster"]
    status_list = [u"wtg", u"rdy", u"ip", u"fin", u"apr", u"pub", u"tmpub", u"rep", u"hld", u"omt"]
    existing_data = []
    if os.path.exists(file_path_xlsx):
        try:
            wb = openpyxl.load_workbook(file_path_xlsx)
            ws = wb.active; rows = list(ws.iter_rows(values_only=True))
            if rows:
                file_headers = [unicode(h) if h else u"" for h in rows[0]]
                for r in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(r):
                        if i < len(file_headers): row_dict[file_headers[i]] = val
                    existing_data.append(row_dict)
        except: pass

    final_rows = []
    processed_assets = set(); sg_back_updates = []
    published_states = ["pub", "tmpub"]

    for row in existing_data:
        asset_name = row.get(u"资产名称")
        if asset_name in sg_map:
            # 1. TexMaster Change detection
            old_tex = row.get("texMaster"); new_tex = sg_map[asset_name].get("texMaster", "")
            if new_tex in ["rev", "rep"]: new_tex = "rep"
            is_tex_break = (old_tex in published_states and new_tex != old_tex)
            row["texMaster"] = new_tex
            
            # 2. RigMaster Change detection
            old_rig = row.get("rigMaster"); new_rig = sg_map[asset_name].get("rigMaster", "")
            if new_rig in ["rev", "rep"]: new_rig = "rep"
            is_rig_break = (old_rig in published_states and new_rig != old_rig)
            row["rigMaster"] = new_rig
            
            # --- [SOP V3.0] Tiered Cascading on Pull ---
            if is_tex_break or new_tex == "rep":
                # Tex changed -> Full Reset
                for ds in ["QC_step_1", "libRig", u"灯光文件制作", "QC_step_2", "libMaster"]:
                    if ds in row:
                        row[ds] = "rep"
                        if ds in ["libRig", "libMaster"]: sg_back_updates.append({"asset_name": asset_name, "task_name": ds, "status": "rep"})
            elif is_rig_break or new_rig == "rep":
                # ONLY Rig changed -> Reset Entry & Master (PROTECT LGT)
                for ds in ["libRig", "libMaster"]:
                    if ds in row:
                        row[ds] = "rep"
                        sg_back_updates.append({"asset_name": asset_name, "task_name": ds, "status": "rep"})
            
            row[u"facialTex"] = sg_map[asset_name].get("facialTex", row.get(u"facialTex", u""))
            processed_assets.add(asset_name)
        final_rows.append(row)
        
    for asset_name in sorted(sg_map.keys()):
        if asset_name not in processed_assets:
            info = sg_map[asset_name]
            new_row = {h: u"" for h in headers}; new_row[u"资产类型"] = info["type"]; new_row[u"资产名称"] = asset_name
            for k in ["texMaster", "rigMaster", "facialTex", "libRig", "libMaster"]:
                v = info.get(k, u""); new_row[k] = "rep" if v in ["rev", "rep"] else v
            final_rows.append(new_row)

    try:
        workbook = xlsxwriter.Workbook(file_path_xlsx); worksheet = workbook.add_worksheet("Assets")
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
        data_format = workbook.add_format({'border': 1})
        for col_num, header in enumerate(headers): worksheet.write(0, col_num, header, header_format)
        for row_num, row_data in enumerate(final_rows):
            excel_row = row_num + 1
            for col_num, header in enumerate(headers):
                val = row_data.get(header, u"")
                worksheet.write(excel_row, col_num, val, data_format)
                if header not in [u"资产类型", u"资产名称", u"制作人员"]:
                    worksheet.data_validation(excel_row, col_num, excel_row, col_num, {'validate': 'list', 'source': status_list})
        worksheet.set_column(0, 0, 10); worksheet.set_column(1, 1, 25); worksheet.set_column(2, len(headers)-1, 15)
        if final_rows: worksheet.autofilter(0, 0, len(final_rows), len(headers) - 1)
        workbook.close()
        if sg_back_updates:
            tmp = os.path.join(os.path.dirname(file_path_xlsx), "tmp_sg_sync_pull.json")
            with io.open(tmp, 'w', encoding='utf-8') as f: f.write(unicode(json.dumps({"project": project_name, "updates": sg_back_updates}, ensure_ascii=False)))
            subprocess.call([BLENDER_PATH, "-b", "-P", SG_UPDATE_SCRIPT, "--", tmp])
    except Exception as e: print("Error: {}".format(e)); sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) > 1: manage_table(sys.argv[1])
